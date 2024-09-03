import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
# from openpyxl.styles import Alignment
# from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
# from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import NamedStyle
from datetime import datetime


# Rutas de los archivos de origen y destino
# "\Users\GMIT\Ingenieria y Gestion Administrativa S.A.S\EQUIPO_GMIT - Documentos\Importe ordenes\RPA Officetrack\Excel_officeTrack\wm_task.xlsx"
archivo_origen = r'C:\Users\GMIT\Ingenieria y Gestion Administrativa S.A.S\EQUIPO_GMIT - Documentos\Importe ordenes\RPA Officetrack\Excel_officeTrack\wm_task.xlsx'
archivo_destino_original = r'C:\Users\GMIT\Ingenieria y Gestion Administrativa S.A.S\EQUIPO_GMIT - Documentos\Importe ordenes\RPA Officetrack\Excel_officeTrack\llenadoDatos.xlsx'
ruta_destino_nueva = r'C:\Users\GMIT\Ingenieria y Gestion Administrativa S.A.S\EQUIPO_GMIT - Documentos\Importe ordenes\RPA Officetrack\Excel_officeTrack\Cambios\nueva_copia.xlsx'

# Cargar el archivo de Excel con la tabla de origen
wb_origen = load_workbook(archivo_origen)
hoja_origen = wb_origen['Page 1']  # Nombre de la hoja

# Cargar el archivo de Excel de destino
wb_destino = load_workbook(archivo_destino_original)
hoja_destino_descarga = wb_destino['Descarga']  # Hoja de destino Descarga
hoja_destino_importe = wb_destino['Importe']  # Hoja de destino Importe

# Copiar los datos desde la fila 2 de la hoja de origen a la hoja de destino Descarga
fila_destino_descarga = 2
for fila in hoja_origen.iter_rows(min_row=2, max_row=hoja_origen.max_row, min_col=1, max_col=12):  # Columna A a L
    for idx, celda in enumerate(fila, start=1):
        hoja_destino_descarga[f"{get_column_letter(idx)}{fila_destino_descarga}"].value = celda.value
    fila_destino_descarga +=1

# Encontrar la primera celda vacía en la fila 2 de la hoja de destino Importe
fila_destino_importe = 2
while hoja_destino_importe[f"B{fila_destino_importe}"].value:  # Mientras la celda no esté vacía
    fila_destino_importe +=1

# Copiar los datos desde la columna A de la hoja de origen a la columna B de la hoja de destino Importe
for fila in hoja_origen.iter_rows(min_row=2, max_row=hoja_origen.max_row, min_col=1, max_col=1):  # Columna A
    for celda_origen in fila:
        celda_destino_importe = hoja_destino_importe.cell(row=fila_destino_importe, column=2)  # Columna B en la fila de destino Importe
        celda_destino_importe.value = celda_origen.value
        fila_destino_importe += 1

fila_destino_importe = 2
for fila in hoja_destino_descarga.iter_rows(min_row=2, max_row=hoja_destino_descarga.max_row, min_col=2, max_col=2):  
    for celda_origen in fila:
        celda_destino_importe = hoja_destino_importe.cell(row=fila_destino_importe, column=3)  
        celda_destino_importe.value = celda_origen.value
        fila_destino_importe += 1

fila_destino_importe = 2
for fila in hoja_destino_descarga.iter_rows(min_row=2, max_row=hoja_destino_descarga.max_row, min_col=2, max_col=1):  
    for celda_origen in fila:
        celda_destino_importe = hoja_destino_importe.cell(row=fila_destino_importe, column=3) 
        celda_destino_importe.value = celda_origen.value
        fila_destino_importe += 1
        fila_destino_importe = 2
        
for fila in hoja_destino_descarga.iter_rows(min_row=2, max_row=hoja_destino_descarga.max_row, min_col=3, max_col=3):  
    for celda_origen in fila:
        celda_destino_importe = hoja_destino_importe.cell(row=fila_destino_importe, column=5) 
        celda_destino_importe.value = celda_origen.value
        fila_destino_importe += 1
        
fila_destino_importe = 2
for fila in hoja_destino_descarga.iter_rows(min_row=2, max_row=hoja_destino_descarga.max_row, min_col=4, max_col=4):  
    for celda_origen in fila:
        # Obtener el valor de la celda origen
        valor_origen = celda_origen.value
        # Verificar si el valor de la celda origen no es None antes de intentar obtener su longitud
        if valor_origen is not None:
            # Truncar el valor a 100 caracteres si es más largo
            valor_destino = valor_origen[:100]
        else:
            valor_destino = None  # O cualquier valor por defecto que desees usar si el valor es None
        # Escribir el valor truncado en la celda destino en la hoja "Importe"
        celda_destino_importe = hoja_destino_importe.cell(row=fila_destino_importe, column=6)  
        celda_destino_importe.value = valor_destino
        fila_destino_importe += 1

fila_destino_importe = 2
for fila in hoja_destino_descarga.iter_rows(min_row=2, max_row=hoja_destino_descarga.max_row, min_col=5, max_col=5):  
    for celda_origen in fila:
        # Obtener el valor de la celda origen
        valor_origen = celda_origen.value
        # Verificar si el valor de la celda origen no es None antes de intentar obtener su longitud
        if valor_origen is not None:
            # Truncar el valor a 100 caracteres si es más largo
            valor_destino = valor_origen[:100]
        else:
            valor_destino = None  # O cualquier valor por defecto que desees usar si el valor es None
        # Escribir el valor truncado en la celda destino en la hoja "Importe"
        celda_destino_importe = hoja_destino_importe.cell(row=fila_destino_importe, column=7)  
        celda_destino_importe.value = valor_destino
        fila_destino_importe += 1



fila_destino_importe = 2
for fila in hoja_destino_descarga.iter_rows(min_row=2, max_row=hoja_destino_descarga.max_row, min_col=6, max_col=6):  
    for celda_origen in fila:
        celda_destino_importe = hoja_destino_importe.cell(row=fila_destino_importe, column=8) 
        celda_destino_importe.value = celda_origen.value
        fila_destino_importe += 1

fila_destino_importe = 2
for fila in hoja_destino_descarga.iter_rows(min_row=2, max_row=hoja_destino_descarga.max_row, min_col=7, max_col=7):  
    for celda_origen in fila:
        celda_destino_importe = hoja_destino_importe.cell(row=fila_destino_importe, column=9)  
        celda_destino_importe.value = celda_origen.value
        fila_destino_importe += 1


fila_destino_importe = 2
for fila in hoja_destino_descarga.iter_rows(min_row=2, max_row=hoja_destino_descarga.max_row, min_col=8, max_col=8):  
    for celda_origen in fila:
        celda_destino_importe = hoja_destino_importe.cell(row=fila_destino_importe, column=10) 
        celda_destino_importe.value = celda_origen.value
        fila_destino_importe += 1
        
fila_destino_importe = 2
for fila in hoja_destino_descarga.iter_rows(min_row=2, max_row=hoja_destino_descarga.max_row, min_col=9, max_col=9): 
    for celda_origen in fila:
        celda_destino_importe = hoja_destino_importe.cell(row=fila_destino_importe, column=11)  
        celda_destino_importe.value = celda_origen.value
        fila_destino_importe += 1


ultima_fila_importe = hoja_destino_importe.max_row
# Escribir "GMIT" en la columna M para cada fila en la hoja "Importe"
for fila in hoja_destino_importe.iter_rows(min_row=2, max_row=hoja_destino_importe.max_row, min_col=2, max_col=2):  # Columna B
    for celda in fila:
        if celda.value:  # Verificar si la celda no está vacía
            celda_m = hoja_destino_importe.cell(row=celda.row, column=13)  # Columna M en la misma fila
            celda_m.value = "GMIT"

        
# Copiar los datos desde la hoja Importe del archivo destino a la misma hoja en el nuevo libro
hoja_nueva_importe = wb_origen.create_sheet(title='Importe')  # Crear una nueva hoja en el nuevo libro
for fila in hoja_destino_importe.iter_rows(min_row=1, max_row=hoja_destino_importe.max_row, min_col=1, max_col=hoja_destino_importe.max_column):
    datos_fila = [celda.value for celda in fila]
    hoja_nueva_importe.append(datos_fila)

# Iterar sobre las filas de la hoja "Importe" para copiar los últimos cinco caracteres de la columna B a la columna A
for fila in hoja_destino_importe.iter_rows(min_row=2, max_row=hoja_destino_importe.max_row, min_col=2, max_col=2):  # Columna B
    for celda in fila:
        if celda.value:  # Verificar si la celda no está vacía
            # Obtener los últimos cinco caracteres del valor en la celda de la columna B
            ultimos_cinco = str(celda.value)[-5:]
            celda_a = hoja_destino_importe.cell(row=celda.row, column=1)  # Columna A en la misma fila
            celda_a.value = ultimos_cinco

from datetime import datetime

# Iterar sobre las filas de la hoja "Importe" para reformatear la fecha y pegarla en la columna H
for fila in hoja_destino_importe.iter_rows(min_row=2, max_row=hoja_destino_importe.max_row, min_col=8, max_col=8):  # Columna F
    for celda in fila:
        fecha_hora = celda.value
        if fecha_hora:
            # Convertir la fecha y hora a un objeto datetime
            fecha_objeto = datetime.strptime(str(fecha_hora), "%Y-%m-%d %H:%M:%S")
            # Reformatear la fecha al formato "dd/mm/yyyy"
            fecha_formateada = fecha_objeto.strftime("%d/%m/%Y")
            # Pegar la fecha formateada en la columna H de la misma fila
            celda_h = hoja_destino_importe.cell(row=celda.row, column=8)  
            celda_h.value = fecha_formateada


for fila in hoja_destino_importe.iter_rows(min_row=2, max_row=hoja_destino_importe.max_row, min_col=9, max_col=9):  
    for celda in fila:
        fecha_hora = celda.value
        if fecha_hora:
            # Convertir la fecha y hora a un objeto datetime
            fecha_objeto = datetime.strptime(str(fecha_hora), "%Y-%m-%d %H:%M:%S")
            # Reformatear la fecha al formato "dd/mm/yyyy"
            fecha_formateada = fecha_objeto.strftime("%d/%m/%Y")
        
            celda_I = hoja_destino_importe.cell(row=celda.row, column=9)  
            celda_I.value = fecha_formateada
            
# Cargar la hoja "Maestro"
hoja_maestro = wb_destino['Maestro']

# Iterar sobre las filas en la hoja de destino "Importe"
for fila_importe in hoja_destino_importe.iter_rows(min_row=2, max_row=hoja_destino_importe.max_row, min_col=5, max_col=5):  # Columna E
    for celda_importe in fila_importe:
        nombre_importe = celda_importe.value
        if nombre_importe is not None:
            nombre_importe = nombre_importe.strip()  # Eliminar espacios en blanco alrededor del nombre
            # Buscar el nombre en la hoja "Maestro" y obtener su cédula
            for fila_maestro in hoja_maestro.iter_rows(min_row=2, max_row=hoja_maestro.max_row, min_col=1, max_col=2):  # Columnas A y B
                nombre_maestro = fila_maestro[0].value
                cedula_maestro = fila_maestro[1].value
                if nombre_maestro is not None and nombre_maestro.strip() == nombre_importe:
                    celda_cedula = hoja_destino_importe.cell(row=celda_importe.row, column=4)  # Columna D en la misma fila
                    celda_cedula.value = cedula_maestro
                    break  # Salir del bucle si se encontró la coincidencia

# Iterar sobre las filas en la hoja de destino "Importe"
for fila_importe in hoja_destino_importe.iter_rows(min_row=2, max_row=hoja_destino_importe.max_row, min_col=11, max_col=11):  # Columna K
    for celda_importe in fila_importe:
        dato_importe = celda_importe.value
        if dato_importe is not None:
            # Buscar el dato en la hoja "Maestro" y obtener el correspondiente en la columna E y F
            for fila_maestro in hoja_maestro.iter_rows(min_row=2, max_row=hoja_maestro.max_row, min_col=4, max_col=5):  # Columnas D y E
                dato_maestro_1 = fila_maestro[0].value
                dato_maestro_2 = fila_maestro[1].value
                if dato_maestro_1 is not None and dato_maestro_1 == dato_importe:
                    celda_destino = hoja_destino_importe.cell(row=celda_importe.row, column=12)  # Columna L en la misma fila
                    celda_destino.value = dato_maestro_2
                    break  # Salir del bucle si se encontró la coincidencia
                
#coger la columna fecha inicio (H) restarla con la fecha fin (I) para calcular las horas y agregarlo en la columna N dd/mm/yyyy
for fila in hoja_destino_importe.iter_rows(min_row=2, max_row=hoja_destino_importe.max_row, min_col=8, max_col=9):  # Columnas H e I
    celda_h = fila[0]
    celda_i = fila[1]
    fecha_inicio = celda_h.value
    fecha_fin = celda_i.value
    if fecha_inicio and fecha_fin:
        # Convertir las fechas de cadena a objetos datetime si es necesario
        if isinstance(fecha_inicio, str):
            fecha_inicio = datetime.strptime(fecha_inicio, "%d/%m/%Y")
        if isinstance(fecha_fin, str):
            fecha_fin = datetime.strptime(fecha_fin, "%d/%m/%Y")
        
        # Calcular la diferencia en días entre la fecha de inicio y la fecha de fin
        diferencia_dias = (fecha_fin - fecha_inicio).days
        
        # Si las fechas son iguales, establecer la diferencia en 540 minutos (9 horas)
        if fecha_inicio.date() == fecha_fin.date():
            diferencia_minutos = 540

        # Contar cada día entre las fechas como 9 horas de trabajo y agregar un día adicional si la diferencia es de al menos 1 día
        else:
            diferencia_dias += 1
            diferencia_minutos = diferencia_dias * 16.5 * 60
        # Escribir el resultado en la columna N
        celda_n = hoja_destino_importe.cell(row=celda_h.row, column=14)  # Columna N en la misma fila
        celda_n.value = diferencia_minutos

            
wb_destino.save(ruta_destino_nueva)

# tabla1= pd.read_excel(ruta_destino_nueva)
# print(tabla1)
#copiar los datos para que los datos creados con formulas se conviertan en

print("El archivo se ha guardado en la nueva ruta:", ruta_destino_nueva)

# Ruta del archivo de Excel original
#"C:\Users\JuanVC\Ingenieria y Gestion Administrativa S.A.S\EQUIPO_GMIT - Documentos\Importe ordenes\RPA Officetrack\Excel_officeTrack\Cambios\nueva_copia.xlsx"
archivo_original = r'C:\Users\GMIT\Ingenieria y Gestion Administrativa S.A.S\EQUIPO_GMIT - Documentos\Importe ordenes\RPA Officetrack\Excel_officeTrack\Cambios\nueva_copia.xlsx'

# Ruta y nombre de la carpeta para el nuevo archivo de Excel
carpeta_nueva = r'C:\Users\GMIT\Ingenieria y Gestion Administrativa S.A.S\EQUIPO_GMIT - Documentos\Importe ordenes\RPA Officetrack\Excel_officeTrack\Cambios'

# Lee la tabla del archivo de Excel original
tabla = pd.read_excel(archivo_original)

# Ruta y nombre del nuevo archivo de Excel
archivo_nuevo = carpeta_nueva + r'\Importe.xlsx'

# Crea un nuevo archivo de Excel y escribe la tabla en él
with pd.ExcelWriter(archivo_nuevo, engine='openpyxl') as writer:
    tabla.to_excel(writer, index=False)
    
print("¡Tabla copiada exitosamente en el nuevo archivo en la ubicación especificada!")